"""
This Script automates the modification of the SNMP ACLs in Cisco routers
and switches to speed up the transition to a new monitoring platform. It
gathers the CDP neighbors information, limiting the user interaction to
only submit its credentials, and a list with the IP addresses of the main
routers.

Note that for the correct execution of this script, the libraries found
below must be installed in the environment where it is being executed.
"""
from getpass import getpass
from netmiko import ConnectHandler
from openpyxl import load_workbook, Workbook
from datetime import date
from os.path import exists
from netmiko.exceptions import NetMikoTimeoutException, NetMikoAuthenticationException

def get_main_routers():
    """
    This function reads a document where the IP addresses
    of the main routers of the sites to be configured are, then,
    stores them in a list.

    Receives:
        None
    Returns:
        routers_list : list
            List of the main routers of the locations
            to be configured.

    """
    router_ips = open("test.txt")
    routers_list = [x.replace('\n', '') for x in router_ips.readlines()]
    return routers_list

def log_creator(file_name):
    """
    This function verifies if an Excel spreadsheet is already present
    to log the changes, and if it is not, it proceeds to create it.

    Receives:
        file_name : str
            Path of the file to be created
    Returns:
        none
        
    """
    if not exists(file_name):
        wb = Workbook()
        page = wb.active
        page.append(['Device', 'Status'])
        wb.save(file_name)

def verify_acls(current_aces):
    """
    This function verifies if the required ACLs are already
    configured on the device. It iterates over the current ACEs,
    if the source ip address is present on the new ACLs, it is removed
    from the corresponding list. 

    At the end, an empty list will mean that the new ACLs are properly
    configured, and the function will return True, otherwise, False will
    be returned. 

    Receives:
        current_aces : list 
            List that contain the current ACEs.
    Returns:
        is_ro_present : boolean
            Indicates if the snmp_ro is configured in
            the device.

        is_rw_present : boolean 
            Indicates if the snmp_rw is configured in
            the device.
    """
    #new ACLs
    snmp_ro = ["10.85.23.25", "10.93.1.56","10.95.1.3", "10.99.72.54", "10.101.72.53", 
    "10.35.72.52", "10.92.72.51", "10.42.72.50", "10.84.75.49", "10.14.32.48", "10.98.72.46",
    "10.91.56.23", "10.93.154.125", "10.31.132.80", "10.12.17.6", "10.95.72.2", "10.75.22.1"]
    snmp_rw = ["10.85.23.25", "10.93.1.56","10.95.1.3", "10.99.72.54", "10.101.72.53", 
    "10.35.72.52", "10.92.72.51", "10.42.72.50", "10.84.75.49", "10.14.32.48", "10.98.72.46",
    "10.91.56.23", "10.93.154.125", "10.31.132.80", "10.12.17.6", "10.95.72.2", "10.75.22.1"]
    for acl in current_aces:
        if acl["acl_name"] == 'SNMP_RO' and acl['src_host'] in snmp_ro:
            snmp_ro.remove(acl['src_host'])
        elif acl["acl_name"] == 'SNMP_RW' and acl['src_host'] in snmp_rw:
            snmp_rw.remove(acl['src_host'])
    is_ro_present = True if len(snmp_ro) == 0 else False
    is_rw_present = True if len(snmp_rw) == 0 else False
    return is_ro_present , is_rw_present 

def get_neighbors(neighbors):
    """
    This function discovers the CDP neighbors of the device being configured (if any),
    and adds them to "neighbors_list".

    Receives:
        neighbors : dict 
            List that contains the CDP neighbors information.
    Returns:
        None
    """
    try:
        for device in neighbors:
            capabilities = (device["capabilities"]).split()
            if "Switch" in capabilities or "Router" in capabilities:
                if device["management_ip"] not in neighbors_list:
                    neighbors_list.append(device["management_ip"])  
    except TypeError:
        print(f"No CDP neighbors were found for {ip}")
    except KeyError as e:
        print(f"Script found the following error {e}. Please contact the network developer")

def log_writer(file_name, entry):
    """
    This function appends to the Excel spreadsheet
    the IP of the device being configured and a message
    that indicates the result of the operation.

    Receives:
        file_name : str 
            Location of the file spreadsheet.

        entry : list
            Information to add in the format ["ip", "message"].  
    Returns:
        None
    """
    wb = load_workbook(file_name)
    page = wb.active
    page.append(entry)
    wb.save(file_name)

if __name__ == "__main__":
    main_routers_list = get_main_routers()
    username = input("Please enter your username: ")
    password = getpass()
    filepath = f'.\\password change log {date.today()}.xlsx'
    log_creator(filepath)
    for router in main_routers_list:
        neighbors_list = []
        neighbors_list.append(router)
        for ip in neighbors_list:
            network_device = {
                "device_type": "cisco_ios",
                "host": ip,
                "username": username,
                "password": password,
            }
            try:
                with ConnectHandler(**network_device) as net_connect:
                    #leverages textfsm to parse the information to a dictionary. 
                    current_acls = net_connect.send_command("show ip access-lists", use_textfsm=True)
                    cdp_neighbors = net_connect.send_command("show cdp neighbors detail", use_textfsm=True)
                    is_ro_config, is_rw_config = verify_acls(current_acls)
                    if not is_ro_config:
                        net_connect.send_config_from_file("ro.txt")
                    if not is_rw_config:
                        net_connect.send_config_from_file("rw.txt")
                    log_writer(filepath, [ip, "Configured"])
                    get_neighbors(cdp_neighbors)
            #The device is unreachable
            except NetMikoTimeoutException:
                log_writer(filepath, [ip, "Failed. A connection could not be established"])
            #The device is reachable but the credentials are incorrect.
            except NetMikoAuthenticationException:
                log_writer(filepath, [ip, "Failed. The device rejected the credentials"])
    print("The configuration has been completed, check the log for more info")